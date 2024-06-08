import http.client
import urllib.parse
import json

from openpyxl import Workbook, load_workbook
import os

import time 
from datetime import datetime

from dotenv import load_dotenv

load_dotenv()

api_key = os.getenv('API_KEY')


dosya_adi = "TRACK & TRACE.xlsx"

mevcut_tarih = datetime.now()

if os.path.exists(dosya_adi):
    workbook = load_workbook(dosya_adi)
    sheet = workbook.active
else:
    workbook = Workbook()
    sheet = workbook.active

max_row = sheet.max_row




for satir_numarasi in range(2, max_row + 1):
    satir = sheet[2]
    cell_value = satir[0].value
    print(cell_value)
    
    params = urllib.parse.urlencode({
        'trackingNumber': cell_value
    })

    headers = {
        'Accept': 'application/json',
        'DHL-API-Key': api_key
    }

    connection = http.client.HTTPSConnection("api-eu.dhl.com")
    connection.request("GET", "/track/shipments?" + params, "", headers)
    response = connection.getresponse()

    status = response.status
    reason = response.reason
    data = json.loads(response.read())

    try:
        current_status = data['shipments'][0]['status']['status']

        status_timestamp = data['shipments'][0]['status']['timestamp']
        print("Status: {} and reason: {}".format(status, reason))
        print("Current Status:", current_status)
        print("Last Status Update Timestamp:", status_timestamp)


        satir[3].value = status_timestamp
        satir[4].value = current_status


        if current_status == "ZU":
            Başarılı_sheet = workbook.get_sheet_by_name("Başarılı")
            Başarılı_sheet.append([cell.value for cell in satir])
            sheet.delete_rows(2, 1)

        elif current_status == "VA":
            Geçen_sheet = workbook.get_sheet_by_name("Henüz çıkmayan")
            Geçen_sheet.append([cell.value for cell in satir])
            sheet.delete_rows(2, 1)
            print("Henüz çıkmamıştır")

        elif current_status == "ZN":
            Geçen_sheet = workbook.get_sheet_by_name("Ulaşılamadı")
            Geçen_sheet.append([cell.value for cell in satir])
            sheet.delete_rows(2, 1)
            print("Ulaşılamadı")
            

        else:
            belirli_tarih = datetime.strptime(status_timestamp, "%Y-%m-%dT%H:%M:%S")
            fark = mevcut_tarih - belirli_tarih
            fark_gun = fark.days

            if fark_gun > 2:
                Geçen_sheet = workbook.get_sheet_by_name("3 Gün Geçen")
                Geçen_sheet.append([cell.value for cell in satir])
                sheet.delete_rows(2, 1)
                print("3 günden fazla zaman geçti.")

                
            else:
                Yolda_sheet = workbook.get_sheet_by_name("Yolda")
                Yolda_sheet.append([cell.value for cell in satir])
                sheet.delete_rows(2, 1)
                print(f"Aradaki fark {fark_gun} gün.")

    except Exception as e:
        print(e)
        hata_sheet = workbook.get_sheet_by_name("Hata")
        hata_sheet.append([cell.value for cell in satir])
        sheet.delete_rows(2, 1)
        
    connection.close()
    print()
    time.sleep(2)

workbook.save(dosya_adi)
